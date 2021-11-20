import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as es
from selenium.webdriver.common.by import By


driver = webdriver.Chrome()
url = r'https://pro.fira.ru/search/#company'
quote_list = []
url1 = 'https://pro.fira.ru/search/companies/card/index.html?code='
url3 = '#!198'
title_list = ['Статус ЕГРЮЛ', 'Телефон', 'Численность сотрудников', 'Вид деятельности', 'Адрес', 'Сайт',
              'E-mail']
stroka_nomer = 2
firma_all_list = []
firma_all_list_now = []


delay = 20
driver.get(url)
input('Нажмите Enter')
myElem = WebDriverWait(driver, delay).until(es.presence_of_element_located((By.CLASS_NAME, 'objbox')))
soup = BeautifulSoup(driver.page_source, 'html.parser')
quotes = soup.find_all(class_='le_active')
for j in quotes:
    firma = str(j)
    firma_list = firma.split(' ')
    firma_list_1 = firma_list[2]
    firma_list_2 = firma_list_1[45:]
    firma_list_3 = firma_list_2[:-1]
    firma_all_list.append(firma_list_3)
firma_all_list = firma_all_list[:20]
firma_all_list_now = firma_all_list_now + firma_all_list


for firm in firma_all_list_now:
    url_now = url1 + firm + url3
    print(url_now)
    delay = 20
    driver.get(url_now)
    soup_w = BeautifulSoup(driver.page_source, 'html.parser')
    error = soup_w.find_all('title')
    error = str(error)
    if error == '[<title>404 Not Found</title>]':
        continue
    else:
        try:
            myElem = WebDriverWait(driver, delay).until(es.presence_of_element_located((By.CLASS_NAME, 'vizitka_text_cut')))
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            quotes = soup.find_all('td', class_='vizitka_table_lh')
            quotes2 = soup.find_all('span', class_='vizitka_text_cut')
            quotes3 = soup.find_all('h1', class_='company_name')
            title_tag = soup.find('meta')
            firma = str(title_tag)
            firma = firma.split(' ')
            if 'выручка' in firma:
                index = firma.index('выручка')
                total = firma[index + 1]
            else:
                total = 0
            quote_list = []
            quote_list2 = []
            quote_list_dict = {}
            for quote in quotes:
                quote = quote.text
                quote = quote.replace('\xa0', ' ')
                quote_list.append(quote)
            for quote in quotes2:
                quote = quote.text
                quote = quote.replace('\xa0', ' ')
                quote_list2.append(quote)
            for i in range(len(quote_list)):
                a = quote_list[i]
                if quote_list[i] in title_list:
                    title = quote_list[i]
                    content = quote_list2[i]
                    quote_list_dict[title] = content
            path = r'D:\\Работа\\Фира_список предприятий\\Общая база.xlsx'
            wb = openpyxl.load_workbook(path)
            sheet = wb['1']
            name = sheet.cell(row=stroka_nomer, column=1)
            region = sheet.cell(row=stroka_nomer, column=11)
            inn = sheet.cell(row=stroka_nomer, column=10)
            for quote in quotes3:
                company = str(quote.text)
                company_list = company.split(',')
                company_name = company_list[0] + company_list[1]
                company_region = company_list[2]
                name.value = company_name
                region.value = company_region
                company_region_inn_ogrn = str(company_list[3])
                company_region_inn = company_region_inn_ogrn.split(' ')
                inn.value = company_region_inn[2]
            d = sheet.cell(row=stroka_nomer, column=9)
            d.value = total
            for i, j in quote_list_dict.items():
                if i == 'Адрес':
                    d = sheet.cell(row=stroka_nomer, column=6)
                    d.value = j
                elif i == 'Статус ЕГРЮЛ':
                    d = sheet.cell(row=stroka_nomer, column=2)
                    d.value = j
                elif i == 'Телефон':
                    d = sheet.cell(row=stroka_nomer, column=3)
                    d.value = j
                elif i == 'Численность сотрудников':
                    d = sheet.cell(row=stroka_nomer, column=4)
                    d.value = j
                elif i == 'Вид деятельности':
                    d = sheet.cell(row=stroka_nomer, column=5)
                    d.value = j
                elif i == 'Сайт':
                    d = sheet.cell(row=stroka_nomer, column=7)
                    d.value = j
                elif i == 'E-mail':
                    d = sheet.cell(row=stroka_nomer, column=8)
                    d.value = j
            wb.save('D:\\Работа\\Фира_список предприятий\\Общая база.xlsx')
            stroka_nomer += 1
        except Exception as ex:
            print(f'Ошибка {ex}')

driver.quit()



